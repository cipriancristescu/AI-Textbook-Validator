#py -3.14 -m streamlit run app.py

import re
import streamlit as st
import fitz  # PyMuPDF
import os
import json
import time
import requests
import pandas as pd
from io import BytesIO
from dotenv import load_dotenv
from typing import Any, Dict, List, Optional, Tuple
import base64

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
load_dotenv()
st.set_page_config(page_title="AI Textbook Auditor", layout="wide", initial_sidebar_state="expanded")

TOGETHER_API_KEY = os.getenv("TOGETHER_API_KEY")
if not TOGETHER_API_KEY:
    st.error("Missing TOGETHER_API_KEY in .env!")
    st.stop()

TOGETHER_BASE_URL = "https://api.together.xyz/v1"

MODEL_VISION = "Qwen/Qwen3.5-397B-A17B"

DEFAULT_CHUNK_SIZE = 5   # pages per vision call
PAGE_DPI           = 150  # lower DPI for multi-page uploads

# ============================================================
# SESSION STATE
# ============================================================
_EMPTY_STATE: Dict[str, Any] = {
    "stage": "upload",
    "pdf_bytes": None,
    "doc_len": 0,
    "fact_report": [],
    "grammar_report": [],
    "fact_debug_log": [],
    "grammar_debug_log": [],
    "audit_ran": False,
    "api_calls_log": [],
}

if "app_state" not in st.session_state:
    st.session_state.app_state = dict(_EMPTY_STATE)

for _k, _v in _EMPTY_STATE.items():
    if _k not in st.session_state.app_state:
        st.session_state.app_state[_k] = _v


# ============================================================
# JSON HELPERS
# ============================================================
def strip_thinking_blocks(text: str) -> str:
    if not isinstance(text, str):
        return text
    return re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()


def _fix_json_trailing_commas(s: str) -> str:
    return re.sub(r",\s*([}\]])", r"\1", s)


def _fix_json_invalid_escapes(s: str) -> str:
    """Fix \z, \a, etc. — invalid JSON escape sequences LLMs sometimes emit."""
    return re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', s)


def _extract_balanced_block(s: str, open_ch: str, close_ch: str) -> Optional[str]:
    start = s.find(open_ch)
    if start == -1:
        return None
    depth, in_string, esc = 0, False, False
    for i in range(start, len(s)):
        ch = s[i]
        if in_string:
            if esc:              esc = False
            elif ch == "\\":    esc = True
            elif ch == '"':     in_string = False
            continue
        if ch == '"':           in_string = True; continue
        if ch == open_ch:       depth += 1
        elif ch == close_ch:
            depth -= 1
            if depth == 0:      return s[start:i + 1]
    return None


def safe_extract_json(text: str) -> Optional[Any]:
    if not isinstance(text, str):
        return None
    text = strip_thinking_blocks(text)
    t = text.strip()
    if t.startswith("```json"): t = t[7:]
    elif t.startswith("```"):   t = t[3:]
    if t.endswith("```"):       t = t[:-3]
    t = t.strip()

    variants = [
        t,
        _fix_json_trailing_commas(t),
        _fix_json_invalid_escapes(t),
        _fix_json_invalid_escapes(_fix_json_trailing_commas(t)),
    ]
    for v in variants:
        try: return json.loads(v)
        except Exception: pass

    for open_ch, close_ch in [("{", "}"), ("[", "]")]:
        block = _extract_balanced_block(t, open_ch, close_ch)
        if block:
            for fix in [
                block,
                _fix_json_trailing_commas(block),
                _fix_json_invalid_escapes(block),
                _fix_json_invalid_escapes(_fix_json_trailing_commas(block)),
            ]:
                try: return json.loads(fix)
                except Exception: pass
    return None


def _coerce_items(data: Any, key: str) -> List[Dict[str, Any]]:
    if isinstance(data, dict) and isinstance(data.get(key), list):
        return [x for x in data[key] if isinstance(x, dict)]
    if isinstance(data, dict):
        for v in data.values():
            if isinstance(v, list):
                return [x for x in v if isinstance(x, dict)]
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    return []


def _dedup_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen: set = set()
    out: List[Dict[str, Any]] = []
    for item in items:
        text    = " ".join(str(item.get("text", "")).lower().split())[:120]
        suggest = " ".join(str(item.get("sugestie", "")).lower().split())[:120]
        key = (str(item.get("pagina", "")), text, suggest)
        if text and key not in seen:
            seen.add(key)
            out.append(item)
    return out


def _append_api_log(entry: Dict[str, Any]) -> None:
    st.session_state.app_state.setdefault("api_calls_log", []).append(entry)


# ============================================================
# TOGETHER AI — VISION 
# ============================================================
def together_vision_multi_json(
    prompt: str,
    base64_images: List[str],
    timeout: int = 2000,
) -> Any:
    """Send multiple page images + a prompt to Qwen Vision, return parsed JSON."""
    content: List[Dict[str, Any]] = [{"type": "text", "text": prompt}]
    for b64 in base64_images:
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/jpeg;base64,{b64}"},
        })

    payload = {
        "model": MODEL_VISION,
        "messages": [{"role": "user", "content": content}],
        "response_format": {"type": "json_object"},
        "max_tokens": 200000,
        "reasoning_effort":"low",
        "reasoning":{"enabled": True}
    }

    log_entry: Dict[str, Any] = {
        "call_type": "vision",
        "model": MODEL_VISION,
        "timestamp": time.strftime("%H:%M:%S"),
        "system_prompt": "(vision — no separate system prompt)",
        "user_text_preview": prompt[:600],
        "user_text_full": prompt,
        "num_images": len(base64_images),
        "attempts": [],
        "final_parsed": None,
        "error": None,
        "ok": False,
    }
    attempt_log: Dict[str, Any] = {
        "attempt": 1,
        "http_status": None,
        "raw_model_output": None,
        "http_response_envelope": None,
        "parsed_ok": False,
        "error": None,
    }

    headers = {
        "Authorization": f"Bearer {TOGETHER_API_KEY}",
        "Content-Type": "application/json",
    }

    try:
        r = requests.post(
            f"{TOGETHER_BASE_URL}/chat/completions",
            headers=headers,
            json=payload,
            timeout=timeout,
        )
        attempt_log["http_status"] = r.status_code
        try:
            envelope = r.json()
            attempt_log["http_response_envelope"] = json.dumps(envelope, ensure_ascii=False, indent=2)
        except Exception:
            attempt_log["http_response_envelope"] = r.text[:4000]

        r.raise_for_status()
        msg = envelope["choices"][0]["message"]
        raw = (msg.get("content") or "").strip()
        reasoning_tokens = envelope.get("usage", {}).get("reasoning_tokens", 0)
        attempt_log["raw_model_output"] = raw
        attempt_log["reasoning_tokens"] = reasoning_tokens

        if not raw:
            raise ValueError(
                f"Model returned empty content. "
                f"reasoning_tokens={reasoning_tokens}, "
                f"completion_tokens={envelope.get('usage',{}).get('completion_tokens',0)}. "
                f"The thinking phase consumed all available tokens — raise max_tokens."
            )

        parsed = safe_extract_json(raw)
        if parsed is None:
            raise ValueError(f"JSON parse failed after all strategies. First 300 chars: {raw[:300]}")

        attempt_log["parsed_ok"] = True
        log_entry["final_parsed"] = parsed
        log_entry["ok"] = True

    except Exception as e:
        attempt_log["error"] = str(e)
        log_entry["error"] = str(e)

    log_entry["attempts"].append(attempt_log)
    _append_api_log(log_entry)

    if log_entry["ok"]:
        return log_entry["final_parsed"]
    return {"_error": log_entry["error"]}


# ============================================================
# PDF → IMAGES
# ============================================================
def pages_to_base64(pdf_bytes: bytes, start_1based: int, end_1based: int, dpi: int = PAGE_DPI) -> List[str]:
    """Render a range of PDF pages to base64 JPEGs."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    for idx in range(start_1based - 1, min(end_1based, len(doc))):
        pix  = doc[idx].get_pixmap(dpi=dpi)
        data = pix.tobytes("jpeg")
        images.append(base64.b64encode(data).decode("utf-8"))
    return images


# ============================================================
# FACT-CHECKING  
# ============================================================
_FACT_PROMPT_TEMPLATE = """Ești auditor tehnic pentru un manual școlar românesc de informatică.
Analizează imaginile paginilor {start}–{end} de mai jos.

Caută EXCLUSIV erori tehnice din categoriile:
1. COD — operatori C/C++ greșiți (cout>>, cin<<), sintaxă imposibilă (Typedef, int:var), bucle for cu variabile inconsistente, literal în condiție (for(i=1;1<n;i++)), void main() / #include<iostream.h> fără avertisment;
2. PSEUDOCOD — cifra 0 înlocuită cu litera o, inconsistențe logice;
3. CONCEPT — stivă descrisă ca FIFO, coadă ca LIFO, definiție fundamental greșită;
4. COMPLEXITATE — complexitate algoritmică evident greșită (O(n) pentru bubble sort etc.);
5. STANDARD — standard C++ greșit prezentat fără contextualizare.

NU raporta: gramatică, diacritice, indentare, stil, using namespace std, bits/stdc++.h, variabile scurte/globale, int main() fără return 0.
Fii conservator — dacă nu ești sigur, nu raporta.

Returnează STRICT JSON (fără text în afara JSON-ului):
{{
  "erori": [
    {{
      "pagina": <numărul paginii din manual>,
      "categorie": "COD|PSEUDOCOD|CONCEPT|COMPLEXITATE|STANDARD",
      "fragment": "<text exact din manual>",
      "corect": "<varianta corectă>",
      "explicatie": "<de ce e greșit>",
      "incredere": <0.0–1.0>
    }}
  ]
}}
Dacă nu găsești erori clare, returnează {{"erori": []}}.
"""


def fact_check_chunk(
    pdf_bytes: bytes,
    start_page: int,
    end_page: int,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    debug: Dict[str, Any] = {
        "pages": f"{start_page}-{end_page}",
        "model": MODEL_VISION,
        "error": None,
        "items_found": 0,
    }

    images = pages_to_base64(pdf_bytes, start_page, end_page)
    if not images:
        debug["error"] = "Nu s-au putut randa paginile"
        return [], debug

    prompt = _FACT_PROMPT_TEMPLATE.format(start=start_page, end=end_page)
    res = together_vision_multi_json(prompt, images)

    if "_error" in res:
        debug["error"] = res["_error"]
        return [], debug

    items = _coerce_items(res, "erori")
    rows: List[Dict[str, Any]] = []
    for item in items:
        fragment = str(item.get("fragment", "")).strip()
        if not fragment:
            continue
        rows.append({
            "validat":   True,
            "pagina":    item.get("pagina", start_page),
            "categorie": str(item.get("categorie", "")).strip(),
            "text":      fragment,
            "sugestie":  str(item.get("corect", "")).strip(),
            "explicatie":str(item.get("explicatie", "")).strip(),
            "incredere": float(item.get("incredere", 0.9) or 0.9),
        })

    debug["items_found"] = len(rows)
    return rows, debug


def fact_check_run_all(
    pdf_bytes: bytes,
    page_start: int,
    page_end: int,
    chunk_size: int,
    status_ref=None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    all_rows:  List[Dict[str, Any]] = []
    all_debug: List[Dict[str, Any]] = []

    for bs in range(page_start, page_end + 1, chunk_size):
        be = min(page_end, bs + chunk_size - 1)
        if status_ref:
            status_ref.write(f"Fact-checking (vision): pages **{bs}–{be}**...")

        rows, dbg = fact_check_chunk(pdf_bytes, bs, be)
        if dbg.get("error") and status_ref:
            status_ref.write(f"Eroare fact-check ({bs}-{be}): {dbg['error']}")
        all_debug.append(dbg)
        all_rows.extend(rows)

    return _dedup_rows(all_rows), all_debug


# ============================================================
# GRAMMAR  
# ============================================================
_GRAMMAR_PROMPT_TEMPLATE = """Rol: Ești filolog român senior, specialist în normele DOOM3 (2021).
Analizează imaginile paginilor {start}–{end} de mai jos.

Identifică erori de limbă română din categoriile:
- O (Ortografie): scriere incorectă, diacritice lipsă sau greșite;
- M (Morfologie): forme flexionare incorecte;
- S (Sintaxă): acord greșit, topică defectuoasă;
- P (Punctuație): virgulă, punct, liniuță incorect folosite;
- D3 (DOOM3 Specific): forme care contravin explicit DOOM3 (2021).

Raportează NUMAI erori clare și localizabile. Ignoră codul-sursă și pseudocodul.
Returnează STRICT JSON cu cel mult {max_errors} elemente:
{{
  "erori": [
    {{
      "pagina": <numărul paginii din manual>,
      "tip": "O|M|S|P|D3",
      "fragment": "<text exact din manual, cu diacriticele corecte>",
      "corect": "<forma corectă>",
      "explicatie": "<justificare scurtă>"
    }}
  ]
}}
Dacă nu găsești erori clare, returnează {{"erori": []}}.
"""

_TIP_MAP = {
    "o": "O", "orthography": "O", "ortografie": "O",
    "m": "M", "morphology": "M", "morfologie": "M",
    "s": "S", "syntax": "S", "sintaxă": "S", "sintaxa": "S",
    "p": "P", "punctuation": "P", "punctuație": "P", "punctuatie": "P",
    "d3": "D3", "doom3": "D3", "doom3 specific": "D3",
}


def grammar_chunk(
    pdf_bytes: bytes,
    start_page: int,
    end_page: int,
    max_errors: int,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    debug: Dict[str, Any] = {
        "pages": f"{start_page}-{end_page}",
        "model": MODEL_VISION,
        "error": None,
        "items_found": 0,
    }

    images = pages_to_base64(pdf_bytes, start_page, end_page)
    if not images:
        debug["error"] = "Nu s-au putut randa paginile"
        return [], debug

    prompt = _GRAMMAR_PROMPT_TEMPLATE.format(start=start_page, end=end_page, max_errors=max_errors)
    res = together_vision_multi_json(prompt, images)

    if "_error" in res:
        debug["error"] = res["_error"]
        return [], debug

    items = _coerce_items(res, "erori")
    rows: List[Dict[str, Any]] = []
    for item in items:
        frag = str(item.get("fragment", "")).strip()
        if not frag:
            continue
        raw_tip = (
            str(item.get("tip", "")).strip()
            or str(item.get("type", "")).strip()
        )
        code = _TIP_MAP.get(raw_tip.lower(), "")
        if not code:
            rl = raw_tip.lower()
            if "ortograf" in rl:                   code = "O"
            elif "morfolog" in rl:                 code = "M"
            elif "sintax" in rl or "syntax" in rl: code = "S"
            elif "punct" in rl:                    code = "P"
            elif "doom" in rl:                     code = "D3"
        rows.append({
            "tip":       f"Grammar ({code})" if code else "Grammar",
            "pagina":    item.get("pagina", start_page),
            "text":      frag,
            "sugestie":  str(item.get("corect", "")).strip(),
            "explicatie":str(item.get("explicatie", "")).strip(),
        })

    debug["items_found"] = len(rows)
    return rows, debug


def grammar_run_all(
    pdf_bytes: bytes,
    page_start: int,
    page_end: int,
    chunk_size: int,
    max_errors_per_chunk: int,
    status_ref=None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    all_rows:  List[Dict[str, Any]] = []
    all_debug: List[Dict[str, Any]] = []

    for bs in range(page_start, page_end + 1, chunk_size):
        be = min(page_end, bs + chunk_size - 1)
        if status_ref:
            status_ref.write(f"Grammar: pages **{bs}–{be}**...")

        rows, dbg = grammar_chunk(pdf_bytes, bs, be, max_errors_per_chunk)
        if dbg.get("error") and status_ref:
            status_ref.write(f"Eroare grammar ({bs}-{be}): {dbg['error']}")
        all_debug.append(dbg)
        all_rows.extend(rows)

    return all_rows, all_debug


# ============================================================
# EXCEL EXPORT
# ============================================================
def _safe_cell(v: Any) -> str:
    if v is None:   return ""
    if isinstance(v, bool): return "TRUE" if v else "FALSE"
    return str(v)


def df_to_excel(df: pd.DataFrame, sheet_name: str = "Report") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.freeze_panes = "A2"

    hfill   = PatternFill("solid", fgColor="1F4E78")
    hfont   = Font(color="FFFFFF", bold=True)
    thin    = Side(style="thin", color="D9D9D9")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = list(df.columns)
    ws.append(cols)
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for _, row in df.iterrows():
        ws.append([row.get(c, "") for c in cols])

    pref = {"validat": 10, "tip": 24, "pagina": 8, "text": 70, "sugestie": 70, "explicatie": 70}
    for ci, col in enumerate(cols, 1):
        letter = get_column_letter(ci)
        maxw = len(col)
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            txt = _safe_cell(cell.value)
            maxw = max(maxw, min(max((len(ln) for ln in (txt.splitlines() or [txt])), default=0), 120))
        ws.column_dimensions[letter].width = pref.get(col.lower()) or min(max(maxw + 2, 12), 80)

    for ri in range(2, ws.max_row + 1):
        ml = 1
        for ci in range(1, ws.max_column + 1):
            ltr = get_column_letter(ci)
            w   = ws.column_dimensions[ltr].width or 15
            txt = _safe_cell(ws.cell(row=ri, column=ci).value)
            lns = txt.splitlines() or [txt]
            u   = max(int(w) - 2, 8)
            ml  = max(ml, sum(max(1, len(ln)//u + (1 if len(ln) % u else 0)) for ln in lns))
        ws.row_dimensions[ri].height = min(max(20 * ml, 20), 180)
    ws.row_dimensions[1].height = 28

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ============================================================
# DEBUG PANEL
# ============================================================
def _full_debug_text(api_log: List[Dict[str, Any]]) -> str:
    lines, sep = [], "=" * 80
    for i, e in enumerate(api_log, 1):
        lines += [
            sep,
            f"CALL #{i}  [{e.get('timestamp','')}]  {e.get('call_type','').upper()}"
            f"  ok={e.get('ok')}  model={e.get('model','')}  images={e.get('num_images',0)}",
            sep,
            "── PROMPT ──",
            e.get("user_text_full", ""),
            "",
        ]
        for att in e.get("attempts", []):
            lines += [
                f"── ATTEMPT {att.get('attempt')}  HTTP {att.get('http_status')} ──",
                "RAW MODEL OUTPUT:",
                att.get("raw_model_output") or "(none)",
            ]
            if att.get("error"):
                lines.append(f"ERROR: {att['error']}")
            if att.get("http_response_envelope"):
                lines += ["HTTP ENVELOPE:", att["http_response_envelope"]]
            lines.append("")
        if e.get("final_parsed") is not None:
            lines += ["── PARSED ──", json.dumps(e["final_parsed"], ensure_ascii=False, indent=2)]
        if e.get("error"):
            lines += [f"── FINAL ERROR ──", e["error"]]
        lines.append("")
    return "\n".join(lines)


def render_debug_downloads(key_suffix: str = "") -> None:
    """Show just the download buttons for the full debug log."""
    api_log = st.session_state.app_state.get("api_calls_log", [])
    if not api_log:
        return
    errors = [e for e in api_log if not e.get("ok")]
    st.caption(f"{len(api_log)} API calls — {len(errors)} failure(s)")
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "Download debug log (.txt)",
            _full_debug_text(api_log).encode("utf-8"),
            "api_debug_log.txt", "text/plain",
            key=f"dl_txt_{key_suffix}",
        )
    with c2:
        st.download_button(
            "Download debug log (.json)",
            json.dumps(api_log, ensure_ascii=False, indent=2).encode("utf-8"),
            "api_debug_log.json", "application/json",
            key=f"dl_json_{key_suffix}",
        )


# ============================================================
# HTML REPORT GENERATION
# ============================================================
_REPORT_CAT_META = {
    # Grammar
    "Grammar (O)":  {"label": "Orthography",  "color": "#2563eb"},
    "Grammar (M)":  {"label": "Morphology",   "color": "#2563eb"},
    "Grammar (S)":  {"label": "Syntax",       "color": "#2563eb"},
    "Grammar (P)":  {"label": "Punctuation",  "color": "#2563eb"},
    "Grammar (D3)": {"label": "DOOM3",        "color": "#2563eb"},
    "Grammar":      {"label": "Grammar",      "color": "#2563eb"},
    # Fact-checking
    "COD":          {"label": "Code Error",   "color": "#dc2626"},
    "PSEUDOCOD":    {"label": "Pseudocode",   "color": "#dc2626"},
    "CONCEPT":      {"label": "Concept",      "color": "#b45309"},
    "COMPLEXITATE": {"label": "Complexity",   "color": "#b45309"},
    "STANDARD":     {"label": "Standard",     "color": "#7c3aed"},
    "DEFINITIE":    {"label": "Definition",   "color": "#b45309"},
    "ALGORITM":     {"label": "Algorithm",    "color": "#b45309"},
}


def generate_html_report(
    fact_report: List[Dict[str, Any]],
    grammar_report: List[Dict[str, Any]],
    textbook_name: str,
) -> str:
    issues: List[Dict[str, Any]] = []
    for item in fact_report:
        issues.append({
            "validat":   bool(item.get("validat", True)),
            "tip":       str(item.get("categorie", "")).strip() or "Technical",
            "capitol":   f"Page {item.get('pagina', '')}",
            "text":      str(item.get("text", "")),
            "sugestie":  str(item.get("sugestie", "")),
            "explicatie":str(item.get("explicatie", "")),
        })
    for item in grammar_report:
        issues.append({
            "validat":   bool(item.get("validat", False)),
            "tip":       str(item.get("tip", "")).strip() or "Grammar",
            "capitol":   f"Page {item.get('pagina', '')}",
            "text":      str(item.get("text", "")),
            "sugestie":  str(item.get("sugestie", "")),
            "explicatie":str(item.get("explicatie", "")),
        })

    data    = {textbook_name: {"filename": textbook_name, "issues": issues}}
    payload = json.dumps(data, ensure_ascii=False)
    cat_json = json.dumps(_REPORT_CAT_META, ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{textbook_name} — Validation Report</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
:root {{
  --bg:#f8fafc; --surface:#ffffff; --border:#e2e8f0;
  --text-primary:#1e293b; --text-secondary:#64748b;
  --accent:#2563eb; --accent-soft:#eff6ff; --radius:8px;
}}
body {{ font-family:'Inter',sans-serif; background:var(--bg); color:var(--text-primary); line-height:1.5; }}
.header {{ background:var(--surface); padding:32px; border-bottom:1px solid var(--border); }}
.header h1 {{ font-size:24px; font-weight:600; margin-bottom:4px; }}
.header p {{ color:var(--text-secondary); font-size:14px; }}
.stats {{ display:flex; gap:16px; padding:24px 32px; flex-wrap:wrap; }}
.stat-card {{ background:var(--surface); border:1px solid var(--border); border-radius:var(--radius); padding:16px; min-width:160px; flex:1; cursor:pointer; transition:border .15s; }}
.stat-card:hover {{ border-color:var(--accent); }}
.stat-card .val {{ font-size:24px; font-weight:600; color:var(--accent); }}
.stat-card .lbl {{ font-size:11px; color:var(--text-secondary); text-transform:uppercase; font-weight:600; letter-spacing:.05em; }}
.controls {{ display:flex; gap:12px; padding:16px 32px; flex-wrap:wrap; align-items:center; position:sticky; top:0; z-index:10; background:var(--bg); border-bottom:1px solid var(--border); }}
.search-box {{ flex:1; min-width:240px; padding:10px 16px; border-radius:var(--radius); border:1px solid var(--border); background:var(--surface); color:var(--text-primary); font-size:14px; outline:none; }}
.search-box:focus {{ border-color:var(--accent); }}
select {{ padding:10px 16px; border-radius:var(--radius); border:1px solid var(--border); background:var(--surface); color:var(--text-primary); font-size:14px; cursor:pointer; outline:none; }}
.issue-list {{ padding:24px 32px; display:flex; flex-direction:column; gap:12px; max-width:1200px; margin:0 auto; }}
.issue-card {{ background:var(--surface); border:1px solid var(--border); border-radius:var(--radius); cursor:pointer; transition:border .15s; }}
.issue-card:hover {{ border-color:var(--accent); }}
.issue-header {{ display:flex; align-items:center; gap:16px; padding:16px 20px; }}
.issue-badge {{ font-size:12px; font-weight:600; color:var(--accent); background:var(--accent-soft); padding:2px 8px; border-radius:4px; white-space:nowrap; }}
.issue-chapter {{ font-size:12px; color:var(--text-secondary); min-width:80px; }}
.issue-text {{ flex:1; font-size:14px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
.issue-arrow {{ color:var(--text-secondary); font-size:12px; transition:transform .2s; }}
.issue-card.open .issue-arrow {{ transform:rotate(90deg); }}
.issue-detail {{ display:none; padding:0 20px 20px; border-top:1px solid var(--border); background:#fafafa; }}
.issue-card.open .issue-detail {{ display:block; }}
.detail-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:20px; margin-top:16px; }}
.detail-label {{ font-size:11px; font-weight:600; text-transform:uppercase; color:var(--text-secondary); margin-bottom:6px; }}
.detail-content {{ font-size:14px; padding:12px; background:var(--surface); border:1px solid var(--border); border-radius:4px; white-space:pre-wrap; }}
.detail-full {{ grid-column:1/span 2; background:#f1f5f9; }}
.empty {{ text-align:center; padding:80px 20px; color:var(--text-secondary); font-size:15px; }}
@media(max-width:800px){{
  .detail-grid {{ grid-template-columns:1fr; }}
  .detail-full {{ grid-column:1; }}
  .stats,.controls {{ padding-left:16px; padding-right:16px; }}
}}
</style>
</head>
<body>
<div class="header">
  <h1>{textbook_name} — Validation Report</h1>
  <p>AI Textbook Auditor &bull; <span id="totalShown"></span></p>
</div>
<div class="stats" id="statsBar"></div>
<div class="controls">
  <input class="search-box" id="search" type="text" placeholder="Search issues...">
  <select id="filterType"><option value="">All Types</option></select>
  <select id="filterPage"><option value="">All Pages</option></select>
</div>
<div class="issue-list" id="issueList"></div>
<script>
const DATA = {payload};
const CAT  = {cat_json};
const issues = DATA[Object.keys(DATA)[0]].issues;

function label(tip) {{ return (CAT[tip]||{{}}).label || tip; }}

function renderStats() {{
  const counts = {{}};
  issues.forEach(i => counts[i.tip] = (counts[i.tip]||0)+1);
  document.getElementById("statsBar").innerHTML = Object.entries(counts)
    .sort((a,b)=>b[1]-a[1])
    .map(([tip,cnt])=>`<div class="stat-card" onclick="setType('${{tip}}')">
      <div class="val">${{cnt}}</div><div class="lbl">${{label(tip)}}</div></div>`).join("");
}}

function renderFilters() {{
  const types = [...new Set(issues.map(i=>i.tip))].sort();
  const pages = [...new Set(issues.map(i=>i.capitol))].sort((a,b)=>{{
    const na=parseInt(a.replace(/\D/g,"")), nb=parseInt(b.replace(/\D/g,""));
    return na-nb;
  }});
  document.getElementById("filterType").innerHTML =
    '<option value="">All Types</option>' + types.map(t=>`<option value="${{t}}">${{label(t)}}</option>`).join("");
  document.getElementById("filterPage").innerHTML =
    '<option value="">All Pages</option>' + pages.map(p=>`<option value="${{p}}">${{p}}</option>`).join("");
}}

function renderIssues() {{
  const q  = document.getElementById("search").value.toLowerCase();
  const ft = document.getElementById("filterType").value;
  const fp = document.getElementById("filterPage").value;
  const filtered = issues.filter(i => {{
    if (ft && i.tip !== ft) return false;
    if (fp && i.capitol !== fp) return false;
    if (q && !(i.text+i.sugestie+i.explicatie).toLowerCase().includes(q)) return false;
    return true;
  }});
  document.getElementById("totalShown").textContent = filtered.length + " / " + issues.length + " issues";
  const list = document.getElementById("issueList");
  if (!filtered.length) {{ list.innerHTML='<div class="empty">No results.</div>'; return; }}
  list.innerHTML = filtered.map(i => `
    <div class="issue-card" onclick="this.classList.toggle('open')">
      <div class="issue-header">
        <span class="issue-badge">${{esc(label(i.tip))}}</span>
        <span class="issue-chapter">${{esc(i.capitol)}}</span>
        <span class="issue-text">${{esc(i.text)}}</span>
        <span class="issue-arrow">›</span>
      </div>
      <div class="issue-detail">
        <div class="detail-grid">
          <div><div class="detail-label">Original Text</div><div class="detail-content">${{esc(i.text)||"—"}}</div></div>
          <div><div class="detail-label">Suggestion</div><div class="detail-content">${{esc(i.sugestie)||"—"}}</div></div>
          <div class="detail-full"><div class="detail-label">Explanation</div><div class="detail-content">${{esc(i.explicatie)||"—"}}</div></div>
        </div>
      </div>
    </div>`).join("");
}}

function esc(s) {{ const d=document.createElement("div"); d.textContent=s||""; return d.innerHTML; }}
function setType(t) {{ const el=document.getElementById("filterType"); el.value=el.value===t?"":t; renderIssues(); }}

document.getElementById("search").addEventListener("input", renderIssues);
document.getElementById("filterType").addEventListener("change", renderIssues);
document.getElementById("filterPage").addEventListener("change", renderIssues);
renderStats(); renderFilters(); renderIssues();
</script>
</body>
</html>"""


# ============================================================
# UI
# ============================================================
st.title("AI Textbook Auditor")
st.caption(f"All analysis via Qwen Vision: `{MODEL_VISION}`")

with st.sidebar:
    st.header("What to run?")
    analysis_mode = st.radio(
        "Select analysis",
        ["Both", "Technical fact-checking", "Grammar"],
        index=0,
    )

    st.divider()
    st.header("Chunking")
    chunk_size     = st.slider("Pages per vision call", 2, 10, DEFAULT_CHUNK_SIZE)
    render_dpi     = st.slider("Render DPI (higher = sharper, slower)", 72, 200, PAGE_DPI)

    st.divider()
    st.header("Grammar")
    max_gram_errors = st.slider("Max grammar errors / chunk", 5, 60, 20)

    st.divider()
    show_api_debug = st.checkbox("Show API debug log", value=True)

    st.divider()
    st.markdown(f"**Model:** `{MODEL_VISION}`")


# 1) UPLOAD 
if st.session_state.app_state["stage"] == "upload":
    with st.container(border=True):
        st.subheader("1) Upload PDF")
        uploaded = st.file_uploader("Select a PDF file", type="pdf")

        if uploaded:
            pdf_bytes = uploaded.getvalue()
            st.session_state.app_state["pdf_bytes"] = pdf_bytes
            doc     = fitz.open(stream=pdf_bytes, filetype="pdf")
            doc_len = len(doc)
            st.session_state.app_state["doc_len"] = doc_len
            st.success(f"Loaded: **{uploaded.name}** — {doc_len} pages")

            c1, c2 = st.columns(2)
            with c1:
                page_start = st.number_input("Start page", min_value=1, max_value=doc_len, value=1)
            with c2:
                page_end = st.number_input("End page", min_value=1, max_value=doc_len, value=doc_len)

            if st.button("Start Audit", type="primary"):
                for k in ["fact_report", "grammar_report", "fact_debug_log",
                          "grammar_debug_log", "api_calls_log"]:
                    st.session_state.app_state[k] = []
                st.session_state.app_state["audit_ran"]    = False
                st.session_state.app_state["page_start"]   = int(page_start)
                st.session_state.app_state["page_end"]     = int(page_end)
                st.session_state.app_state["pdf_name"]     = uploaded.name
                st.session_state.app_state["stage"]        = "analyze"
                st.rerun()


# 2) ANALYZE + REPORT
elif st.session_state.app_state["stage"] == "analyze":
    pdf_bytes  = st.session_state.app_state["pdf_bytes"]
    doc_len    = st.session_state.app_state["doc_len"]
    audit_ran  = st.session_state.app_state.get("audit_ran", False)
    page_start = st.session_state.app_state.get("page_start", 1)
    page_end   = st.session_state.app_state.get("page_end", doc_len)

    st.subheader(f"Analysis — pages {page_start}–{page_end}")

    if not audit_ran:
        status = st.status("Running audit...", expanded=True)
        st.session_state.app_state["api_calls_log"] = []

        fact_rows:    List[Dict[str, Any]] = []
        grammar_rows: List[Dict[str, Any]] = []
        fact_debug:   List[Dict[str, Any]] = []
        grammar_debug:List[Dict[str, Any]] = []

        if analysis_mode in ("Both", "Technical fact-checking"):
            status.write(f"Fact-checking pages **{page_start}–{page_end}** via vision...")
            fact_rows, fact_debug = fact_check_run_all(
                pdf_bytes, page_start, page_end, chunk_size,
                status_ref=status,
            )

        if analysis_mode in ("Both", "Grammar"):
            status.write(f"Grammar check pages **{page_start}–{page_end}**...")
            grammar_rows, grammar_debug = grammar_run_all(
                pdf_bytes, page_start, page_end, chunk_size, max_gram_errors,
                status_ref=status,
            )

        st.session_state.app_state["fact_report"]     = fact_rows
        st.session_state.app_state["grammar_report"]  = grammar_rows
        st.session_state.app_state["fact_debug_log"]  = fact_debug
        st.session_state.app_state["grammar_debug_log"] = grammar_debug
        st.session_state.app_state["audit_ran"]       = True
        status.update(label="Analysis complete!", state="complete", expanded=False)
        st.rerun()

    # Results
    fact_report    = st.session_state.app_state.get("fact_report", [])
    grammar_report = st.session_state.app_state.get("grammar_report", [])

    if not fact_report and not grammar_report:
        st.warning("Audit ran but no issues found. Check the API Debug Log below.")

    # Fact-checking table
    if analysis_mode in ("Both", "Technical fact-checking"):
        st.markdown("### Technical Fact-checking")
        if not fact_report:
            st.info("No technical errors found.")
        else:
            st.success(f"{len(fact_report)} technical error(s) found")
            df_f = pd.DataFrame(fact_report)
            for col, default in [
                ("validat", True), ("pagina", ""), ("categorie", ""),
                ("text", ""), ("sugestie", ""), ("explicatie", ""), ("incredere", 0.0)
            ]:
                if col not in df_f.columns:
                    df_f[col] = default
            df_f = df_f[["validat", "pagina", "categorie", "text", "sugestie", "explicatie", "incredere"]]

            counts = df_f["categorie"].replace("", "Necategorizat").value_counts()
            st.caption("By category: " + ", ".join(f"{tip}: {cnt}" for tip, cnt in counts.items()))

            edited_f = st.data_editor(
                df_f,
                column_config={
                    "validat":   st.column_config.CheckboxColumn("Keep?", default=True),
                    "pagina":    st.column_config.NumberColumn("Page", format="%d"),
                    "categorie": "Category",
                    "text":      "Fragment",
                    "sugestie":  "Correction",
                    "explicatie":"Explanation",
                    "incredere": st.column_config.NumberColumn("Confidence", format="%.2f"),
                },
                width="stretch", height=420, key="fact_editor",
            )
            valid_f = edited_f[edited_f["validat"] == True].copy()
            c1, c2 = st.columns([1.2, 1.6])
            with c1:
                if not valid_f.empty:
                    st.download_button("Download CSV (fact-checking)",
                        valid_f.to_csv(index=False).encode("utf-8"),
                        "fact_checking.csv", "text/csv")
            with c2:
                if not valid_f.empty:
                    st.download_button("Download Excel (fact-checking)",
                        df_to_excel(valid_f, "Fact checking"),
                        "fact_checking.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if analysis_mode == "Both":
        st.divider()

    # Grammar table 
    if analysis_mode in ("Both", "Grammar"):
        st.markdown("### Grammar")
        if not grammar_report:
            st.info("No grammar errors found.")
        else:
            st.success(f"{len(grammar_report)} grammar issue(s) found")
            df_g = pd.DataFrame(grammar_report)
            for col, default in [
                ("validat", False), ("tip", ""), ("pagina", ""),
                ("text", ""), ("sugestie", ""), ("explicatie", "")
            ]:
                if col not in df_g.columns:
                    df_g[col] = default
            df_g = df_g[["validat", "tip", "pagina", "text", "sugestie", "explicatie"]]

            counts_g = df_g["tip"].value_counts()
            st.caption("By type: " + ", ".join(f"{tip}: {cnt}" for tip, cnt in counts_g.items()))

            edited_g = st.data_editor(
                df_g,
                column_config={
                    "validat":   st.column_config.CheckboxColumn("Keep?", default=False),
                    "tip":       "Type",
                    "pagina":    st.column_config.NumberColumn("Page", format="%d"),
                    "text":      "Fragment",
                    "sugestie":  "Correction",
                    "explicatie":"Explanation",
                },
                width="stretch", height=420, key="grammar_editor",
            )
            valid_g = edited_g[edited_g["validat"] == True].copy()
            c1, c2 = st.columns([1.2, 1.6])
            with c1:
                if not valid_g.empty:
                    st.download_button("Download CSV (grammar)",
                        valid_g.to_csv(index=False).encode("utf-8"),
                        "grammar.csv", "text/csv")
            with c2:
                if not valid_g.empty:
                    st.download_button("Download Excel (grammar)",
                        df_to_excel(valid_g, "Grammar"),
                        "grammar.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.divider()

    # HTML Report 
    if fact_report or grammar_report:
        pdf_name = st.session_state.app_state.get("pdf_name", "textbook")
        report_name = re.sub(r"\.[^.]+$", "", pdf_name)
        html_bytes = generate_html_report(fact_report, grammar_report, report_name).encode("utf-8")
        st.download_button(
            "Generate & Download HTML Report",
            html_bytes,
            f"{report_name}_report.html",
            "text/html",
        )

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Re-run analysis"):
            for k in ["fact_report", "grammar_report", "fact_debug_log",
                      "grammar_debug_log", "api_calls_log"]:
                st.session_state.app_state[k] = []
            st.session_state.app_state["audit_ran"] = False
            st.rerun()
    with c2:
        if st.button("Upload a different document"):
            st.session_state.app_state = dict(_EMPTY_STATE)
            st.rerun()

    if show_api_debug:
        render_debug_downloads("analyze")
